import gym
from matplotlib import pyplot as plt
from torch.autograd import Variable
from collections import namedtuple
import torch
import torch.nn as nn
import random
import time
import numpy as np
import torch.nn.functional as F
from highway_env import utils
import highway_env
env = gym.make('highway-v0').unwrapped
import sys

import warnings
warnings.filterwarnings("ignore")
import openpyxl as op

from DQN_HEAD.DQN_HEAD import get_config
parser = get_config()
config = parser.parse_args()
###################   Basic   ##########################
# env.config["right_lane_reward"] = 1  # 0.1
env.config["lanes_count"] = config.lanes_count # 4
env.config["show_trajectories"] =  config.show_trajectories
env.config["screen_height"] = config.screen_height  # 150
env.config["screen_width"] = config.screen_width  # 600
env.config["reward_speed_range"] = config.reward_speed_range     # SV静止时用  设定速度范围 [2, 4, 6, 8]
env.config["punish_speed_range"] = config.punish_speed_range

env.config["policy_frequency"] = config.policy_frequency  #  abstract： frames = int(self.config["simulation_frequency"] // self.config["policy_frequency"])
#############  DQN 调参  ###########################
env.config["simulation_frequency"] = config.simulation_frequency  # 决定dt=1/"simulation_frequency"；即：一次step()，系统做出几次控制循环。还和 controller 比例控制 TAU_ACC = 2 有关，一并调整
# 太大，控制结束了，但是下一个动作 a 还没有来。太小的话，控制还没完，下一个动作 a 就来了# KP_A = 1 / 0.375 = 8/3   dt * KP_A = 8/3 * 1/4 = 2/3
env.config["vehicles_count"] = config.vehicles_count #  旁车，数据越大，训练频率变慢，显示越稳定. 每次控制循环实施的次数："vehicles_count"+1
env.config["duration"] = config.duration  # 40  决定了在不 crashed 的前提下，能有多少个 a 和 env.step(a)。直到steps计数到["duration"]，触发done，系统复位。

env.config["collision_reward"] = config.collision_reward  # -1
env.config["high_speed_reward"] = config.high_speed_reward  # 0.4
env.config["punish_speed_reward"] = config.punish_speed_reward
env.config["wrong_act_reward"] = config.wrong_act_reward  #
env.config["correct_act_reward"] = config.correct_act_reward

T_N_update_FQ = config.T_N_update_FQ  # 更新目标网络频率
BATCH_SIZE = config.BATCH_SIZE  # 提取批次数量
MEMORY_CAPACITY = config.MEMORY_CAP * BATCH_SIZE        # 池子容量
total_train_times = config.total_train_t * MEMORY_CAPACITY    # 训练次数

class D3QNnet(nn.Module):
    def __init__(self):
        super(D3QNnet, self).__init__()
        self.V_1 = nn.Linear(config.input_net, config.neure)   #
        self.V_1.weight.data.normal_(0, 0.1)
        self.V_OUT = nn.Linear(config.neure, 1)    # neure = 1024。 输出只有一个 Value
        self.V_OUT.weight.data.normal_(0, 0.1)

        self.A_1 = nn.Linear(config.input_net, config.neure)   #
        self.A_1.weight.data.normal_(0, 0.1)
        self.A_OUT = nn.Linear(config.neure, config.output_net)    # neure = 1024。 输出只有一个 Value
        self.A_OUT.weight.data.normal_(0, 0.1)

    def forward(self, s):
        s = torch.FloatTensor(s)  #
        s = s.view(s.size(0), 1, 20)  # size(0)返回s的行数，view把s拉成1个1行 neure 列才能送入网络

        s01 = self.V_1(s)
        V = F.relu(s01)                     # out = torch.sigmoid(s02)
        V_OT = self.V_OUT(V)
        # print('      V_out:',  V_OT)

        s02 = self.A_1(s)
        A = F.relu(s02)                     # out = torch.sigmoid(s02)
        A_OT = self.A_OUT(A)
        # print('       A_OT:',  A_OT)
        # print('    mean(A_OT):',  torch.mean(A_OT.squeeze(1),1))
        # print(' mean(A_OT)_un:',  torch.mean(A_OT.squeeze(1),1).unsqueeze(1))
        # print('      A-MEAN_A:',  A_OT.squeeze(1) - torch.mean(A_OT.squeeze(1),1).unsqueeze(1))

        Q = V_OT.squeeze(1) + (A_OT.squeeze(1) - torch.mean(A_OT.squeeze(1),1).unsqueeze(1))

        # print('Q:', Q)

        return Q  # 1*5
class DDDQN(object):
    def __init__(self):
        self.eval_net, self.target_net, = D3QNnet(), D3QNnet()
        self.learn_step_counter = 0
        self.memory = []
        self.position = 0
        self.capacity = MEMORY_CAPACITY
        self.optimizer = torch.optim.Adam(self.eval_net.parameters(), lr=config.LR)
        self.loss_func = nn.MSELoss()
        self.index = 0
        for param, target_param in zip(self.eval_net.parameters(), self.target_net.parameters()):
            target_param.data.copy_(param.data)
        self.ZHEKOU = 0.01

    def choose_action(self, s, e):  # s: 5*5矩阵
        x = np.expand_dims(s, axis=0)  # 增加了一维，多了一对中括号
        if np.random.uniform() < 1-e:  # - e
            actions_value = self.eval_net.forward(x)
            action = torch.max(actions_value, -1)[1].data.numpy()  # '-1'可视为‘1’，输出行最大值的索引，如果没有[1]就输出 最大值 和 索引
            action = action.max()
        else:
            action = np.random.randint(0, 5)
        return action

    def push_memory(self, s, a, r, s_):
        if len(self.memory) < self.capacity:   # MEMORY_CAPACITY = self.capacity
            self.memory.append(None)
        self.index = self.position % self.capacity
        self.memory[self.index] = Transition(   torch.unsqueeze(torch.FloatTensor(s), 0),
                                                torch.unsqueeze(torch.FloatTensor(s_), 0), \
                                                torch.from_numpy(np.array([a])),
                                                torch.from_numpy(np.array([r], dtype='float32')))
        self.position += 1

    def learn(self):
        # if self.learn_step_counter % TARGET_NETWORK_REPLACE_FREQ == 0:
        #     self.target_net.load_state_dict(self.eval_net.state_dict())
        transitions = random.sample(self.memory, BATCH_SIZE)
        batch = Transition(*zip(*transitions))  # 按照类型重建元组，s, s_, a, r, 每个元素里面分别有 BATCH_SIZE 个
        b_s = Variable(torch.cat(batch.state))
        b_s_ = Variable(torch.cat(batch.next_state))
        b_a = Variable(torch.cat(batch.action))
        b_r = Variable(torch.cat(batch.reward))

        q_eval = self.eval_net.forward(b_s).squeeze(1).gather(1, b_a.unsqueeze(1).to(torch.int64))

        max_a_index = self.eval_net.forward(b_s_).squeeze(1).max(1)[1].detach()  # .to(torch.int64)
        # print(' max_a_index :', max_a_index )
        q_next = self.target_net.forward(b_s_).squeeze(1).gather(1, max_a_index.unsqueeze(1)).detach()
        # print('q_next:', q_next)

        q_target = b_r + config.GAMMA * q_next.view(BATCH_SIZE, 1).t()

        loss = self.loss_func(q_eval, q_target.t())
        self.optimizer.zero_grad()  # reset the gradient to zero
        loss.backward()
        self.optimizer.step()  # execute back propagation for one step

        if self.learn_step_counter % T_N_update_FQ == 0:  # 一开始触发，然后每100步触发
            for param, target_param in zip(self.eval_net.parameters(), self.target_net.parameters()):
                target_param.data.copy_((1 - self.ZHEKOU) * target_param.data + self.ZHEKOU * param.data)

        self.learn_step_counter += 1

        return loss.detach().numpy()

dddqn = DDDQN()
Transition = namedtuple('Transition', ('state', 'next_state', 'action', 'reward'))  # namedtuple()是继承自tuple的子类，namedtuple()方法能够创建一个和tuple类似的对象，而且对象拥有可访问的属性
# pprint.pprint(env.config)

wb = op.Workbook()  # 创建工作薄对象
ws = wb['Sheet']  # 创建子表
#------------------------------------------ AR 截图上限----------------------------------------------
in_pool_TR = 1  # 进池子标志
# in_pool_TR = 0  # 不进池子
while True:  # 原始程序9900100
    done = False
    s = env.reset()   # 5*5, r, false, {speed, crashed, action, cost}
    config.episode += 1
    config.S_tm_ep = time.time()

    while not done:
        # env.render()  # 显示动画，没有必要      # e = np.exp(-total_env_times / MEMORY_CAPACITY)  # 随机选择action的概率，随着训练次数增多逐渐降低，但是速度太慢，影响池子数据质量。
        a = dddqn.choose_action(s, config.e)  # 对应 ACTIONS_ALL = { 0: 'LANE_LEFT', 1: 'IDLE', 2: 'LANE_RIGHT',  3: 'FASTER',  4: 'SLOWER'   }
        config.total_step_times += 1

        if (config.start_train_TR == 0):  # AR 训练前
            if (config.lane_old == 0 and a == 0) or (config.lane_old == env.config["lanes_count"]-1 and a == 2):  # 如果出现不合理动作，触发惩罚标记 WRO_ACT_TR
                # WRONG_ACT_CNT_bf_train += 1
                WRO_ACT_TR = 1  # 此处标记是为了触发惩罚。
            # else:  # 合理动作
            #     correct_act_cnt_bf_train += 1  # 伪正确

            s_, r, done, info, SUCCESS = env.step(a)  # 错误动作也进入训练，带惩罚。 success 表示成功完成全部轮次。env.config["duration"] = 90
            config.lane_old = r[1]


            scaled_speed = utils.lmap(info['speed'], env.config["reward_speed_range"], [0, 1])  #
            h_spd = env.config["high_speed_reward"] * np.clip(scaled_speed, 0, 1)
            config.reward = r[0] + h_spd
            config.reward = utils.lmap(config.reward,
                                [ env.config["collision_reward"], env.config["high_speed_reward"] ],
                                [0, 1])

            if not done:  # 合理还正确。     训练前 填池子，训练开始后池子数量不增加
                dddqn.push_memory(s, a, config.reward, s_)
                config.memory_cnt += 1
            else:  # 合理不正确。      动作引起的下一时刻撞车动作
                if in_pool_TR == 1:  # 进池子标志，# 正确动作引起的下一时刻撞车动作，进了池子后可以训练网络不输出类似有撞车可能的动作。
                    dddqn.push_memory(s, a, config.reward, s_)  # 错误动作进池子
                    config.memory_cnt += 1  # 错误动作进池子

                config.E_tm_ep = time.time()
                config.tm_ep = config.E_tm_ep - config.S_tm_ep
                config.tm_ep_array.append(config.tm_ep)
                config.ntm += 1
        #========================# 训练后 #===========================#
        elif (config.start_train_TR == 1):  # 训练后
            if (config.lane_old == 0 and a == 0) or (config.lane_old == env.config["lanes_count"]-1 and a == 2):  # 如果出现不合理动作，触发惩罚标记 WRO_ACT_TR
                config.WRONG_ACT_CNT_in_train += 1
            else:
                config.correct_act_cnt_in_train += 1
                config.sum_RA_episode +=1

            s_, r, done, info, SUCCESS = env.step(a)  # info里面含有 speed， crushed，action 的信息。 如：if info['crashed'] == True :
            config.lane_old = r[1]

            scaled_speed = utils.lmap(info['speed'], env.config["reward_speed_range"], [0, 1])  #
            h_spd = env.config["high_speed_reward"] * np.clip(scaled_speed, 0, 1)
            config.reward = r[0] + h_spd
            config.reward = utils.lmap(config.reward, [ env.config["collision_reward"], env.config["high_speed_reward"] ],
                                [0, 1])

            config.reward_array.append(config.reward)  # 积累奖励

            if np.all(s_[1] == [0, 0, 0, 0]) & np.all(s_[2] == [0, 0, 0, 0]):  # 只记录完成避障所用时间
                config.E_tm_ep = time.time()
                config.tm_ep = config.E_tm_ep - config.S_tm_ep
                config.tm_ep_array.append(config.tm_ep)
                config.ntm += 1
                done = True  # 完成避障

            if info['speed']<12:
                config.speed_array.append((info['speed']-3)*2+20)
            else:
                config.speed_array.append(info['speed'])

            if not done:  # 训练前 填池子，训练开始后池子数量不增加
                dddqn.push_memory(s, a, config.reward, s_)
            else:  # 有可能成功避障，由 duration 引发 done
                if in_pool_TR == 1:  # 错误动作进池子标志:撞车的动作和结果
                    dddqn.push_memory(s, a, config.reward, s_)  # 错误动作进池子
                config.done_in_train += 1  # 正确和错误的动作都会引发下一个 done ！！！！！！！

                config.sum_reward_episode = np.sum(config.reward_array)  # 每一轮奖励和
                config.reward_sum_in_train += config.sum_reward_episode  # 全部训练奖励和 最后输出数据用
                config.sum_reward_episode_array.append(config.sum_reward_episode)  # 构成奖励数组 绘图用
                config.nreward += 1
                config.reward_array = []

                config.avg_speed_episode = np.mean(config.speed_array)  # 轮次平均速度
                config.avg_speed_episode_array.append(config.avg_speed_episode)  # 所有轮次平均速度 绘图用
                config.nspeed += 1
                config.speed_array = []

                # config.sum_RA_episode_array.append(config.sum_RA_episode)  # 每一轮的RA数量进入数组
                # config.nRA += 1
                # config.sum_RA_episode = 0

                # avg_loss_episode = np.mean(loss_array)
                # avg_loss_episode_array.append(avg_loss_episode)    # 所有轮次平均损失 绘图用
                # nloss += 1
                # loss_array = []

                # config.E_tm_ep = time.time()
                # config.tm_ep = config.E_tm_ep - config.S_tm_ep
                # config.tm_ep_array.append(config.tm_ep)
                # config.ntm += 1

                if np.all(s_[1] == [0,0,0,0]) & np.all(s_[2] == [0,0,0,0]):
                    config.success_done += 1  # env.commom.abstract 中定义
                    if (config.success_done == 1):
                        end_time = time.time()  # 计算第一次成功避障所用时间
                        config.First_finish_time = end_time - start_time

# ----------------------------------------截图下限-----------------------------------------------------------------------
        s = s_
        print('*  DuDQN-SPD total_step_times:', config.total_step_times)
################################  开始训练  ###################################################
        if dddqn.position >= MEMORY_CAPACITY:  # 池子满了，开始训练
            config.start_train_TR = 1
            config.trained_times += 1
            if (config.trained_times ==1 ):
                start_time = time.time()   # 训练时间计时
            loss = dddqn.learn()
            config.loss_array.append(loss)

            e = 1/(config.trained_times/50 + 10)

            if (config.trained_times == total_train_times + 1 ):  # 全部训练完成，显示结果。跳出循环
                end_time = time.time()
                total_time = end_time - start_time

                # plt.figure()
                # plt.subplot(3, 1, 1)
                # plt.plot(avg_loss_episode_array)
                # plt.xlabel('D3QN AR trained episodes')
                # plt.ylabel('avg_loss')
                # plt.subplot(3, 1, 2)
                # plt.plot(sum_reward_episode_array)
                # plt.xlabel('D3QN AR trained episodes ')
                # plt.ylabel('sum_reward_episode')
                # plt.subplot(3, 1, 3)
                # plt.plot(avg_speed_episode_array)
                # plt.xlabel('D3QN AR trained episodes ')
                # plt.ylabel('avg_speed_episode')
                # plt.tight_layout()

                print('------------------------------------', )
                print('==== DuDQN_SPD Results ===============', )
                print('       memory_cnt:', config.memory_cnt)
                print('      训练次数:', config.trained_times - 1)
                print('   总体训练时间:', np.round(total_time, 0)  )
                print('   首次成功用时:', np.round(config.First_finish_time, 0)  )
                print('训练后合理动作数:', config.correct_act_cnt_in_train)
                print('   合理动作比例:', np.round(100 * config.correct_act_cnt_in_train / (config.trained_times - 1), 2), '%')
                print('   训练奖励总和:', np.round(config.reward_sum_in_train, 0))
                print('     训练后轮次：', config.done_in_train)
                print('   成功避障次数:', config.success_done)
                print('        成功率:', np.round(100 * config.success_done/(config.done_in_train+0.1), 2), '%')
                print('            e:', np.round(e, 6))
                print('-------------------------------', )
                print('Waiting for writing excel......')

                # for i in range(nloss):
                #     ws.cell(row=i + 1, column=3).value = float(avg_loss_episode_array[i])  # 在第1行第1列写入LOSS
                #     wb.save('RRDuDQN-LOSS.xlsx')  # 保存excel表

                for gg in range(config.nspeed):
                    ws.cell(row=gg + 1, column=1).value = float(config.avg_speed_episode_array[gg])  #S
                    wb.save('04-DuDQN-SPD-speed.xlsx')  # 保存excel表

                for hh in range(config.nreward):
                    ws.cell(row=hh + 1, column=1).value = float(config.sum_reward_episode_array[hh])  #
                    wb.save('04-DuDQN-SPD-reward.xlsx')  # 保存excel表

                # for l in range(config.nRA):
                #     ws.cell(row=l + 1, column=1).value = float(config.sum_RA_episode_array[l])  #
                #     wb.save('04-DuDQN-SPD-RA.xlsx')  # 保存excel表

                for ii in range(config.ntm):
                    ws.cell(row=ii + 1, column=1).value = float(config.tm_ep_array[ii])  #
                    wb.save('04-DuDQN-SPD-tmep.xlsx')  #

                print('Writing excel is finished.')

                plt.show()
                # break
                sys.exit()